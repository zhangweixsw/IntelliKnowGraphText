import os
import threading
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from PyQt6.QtWidgets import (
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QFileDialog,
    QProgressBar,
    QSplitter,
    QGroupBox,
    QMessageBox,
    QStatusBar,
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt6.QtGui import QFont, QColor, QPalette

from core.document_processor import DocumentProcessor, ProcessingResult
from data.database import db, Document
from utils.config import config
from utils.model_detector import ModelDetector
from ui.settings_dialog import SettingsDialog


class WorkerSignals(QObject):
    progress = pyqtSignal(str)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.processor = DocumentProcessor()
        self.current_files: List[str] = []
        self.results: List[ProcessingResult] = []
        self.worker_thread: Optional[threading.Thread] = None
        self.worker_signals = WorkerSignals()
        self.init_ui()
        self._connect_signals()
        # 移除启动时的自动模型检测 - 改为延迟检测
        # 用户点击设置或开始处理时才检测

    def _connect_signals(self):
        self.worker_signals.progress.connect(self.on_processing_progress)
        self.worker_signals.finished.connect(self.on_processing_finished)

    def init_ui(self):
        self.setWindowTitle("猫头羊zhangweixsw - 智能文档识别系统")
        self.setGeometry(100, 100, 1200, 800)

        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#1e1e1e"))
        palette.setColor(QPalette.ColorRole.WindowText, QColor("#d4d4d4"))
        palette.setColor(QPalette.ColorRole.Base, QColor("#252526"))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#2d2d2d"))
        palette.setColor(QPalette.ColorRole.ToolTipBase, QColor("#2d2d2d"))
        palette.setColor(QPalette.ColorRole.Text, QColor("#d4d4d4"))
        palette.setColor(QPalette.ColorRole.Button, QColor("#3c3c3c"))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor("#d4d4d4"))
        palette.setColor(QPalette.ColorRole.Highlight, QColor("#007acc"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
        self.setPalette(palette)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)

        title_label = QLabel("猫头羊zhangweixsw - 智能文档识别系统")
        title_font = QFont("Microsoft YaHei", 14, QFont.Weight.Bold)
        title_label.setFont(title_font)
        title_label.setStyleSheet("color: #4ec9b0; padding: 5px;")
        main_layout.addWidget(title_label)

        toolbar = self._create_toolbar()
        main_layout.addWidget(toolbar)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(splitter)

        left_panel = self._create_left_panel()
        splitter.addWidget(left_panel)

        right_panel = self._create_right_panel()
        splitter.addWidget(right_panel)

        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)

        status_bar = QStatusBar()
        self.setStatusBar(status_bar)
        self.status_label = QLabel("就绪")
        status_bar.addWidget(self.status_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        status_bar.addPermanentWidget(self.progress_bar)

    def _create_toolbar(self) -> QWidget:
        toolbar = QWidget()
        layout = QHBoxLayout(toolbar)
        layout.setContentsMargins(0, 0, 0, 0)

        self.btn_select_file = QPushButton("📁 选择文件")
        self.btn_select_file.clicked.connect(self.select_files)
        layout.addWidget(self.btn_select_file)

        self.btn_select_dir = QPushButton("📂 选择文件夹")
        self.btn_select_dir.clicked.connect(self.select_directory)
        layout.addWidget(self.btn_select_dir)

        self.btn_start = QPushButton("▶ 开始处理")
        self.btn_start.clicked.connect(self.start_processing)
        self.btn_start.setEnabled(False)
        layout.addWidget(self.btn_start)

        self.btn_stop = QPushButton("⏹ 停止")
        self.btn_stop.clicked.connect(self.stop_processing)
        self.btn_stop.setEnabled(False)
        layout.addWidget(self.btn_stop)

        layout.addSpacing(20)

        self.btn_export_md = QPushButton("📝 导出MD")
        self.btn_export_md.clicked.connect(self.export_md)
        layout.addWidget(self.btn_export_md)

        self.btn_export_excel = QPushButton("📊 导出Excel")
        self.btn_export_excel.clicked.connect(self.export_excel)
        layout.addWidget(self.btn_export_excel)

        layout.addStretch()

        self.btn_settings = QPushButton("⚙ 设置")
        self.btn_settings.clicked.connect(self.open_settings)
        layout.addWidget(self.btn_settings)

        for btn in [self.btn_select_file, self.btn_select_dir, self.btn_start,
                    self.btn_stop, self.btn_export_md, self.btn_export_excel,
                    self.btn_settings]:
            btn.setMinimumHeight(36)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #3c3c3c;
                    border: 1px solid #555;
                    border-radius: 4px;
                    padding: 5px 15px;
                    color: #d4d4d4;
                }
                QPushButton:hover {
                    background-color: #4a4a4a;
                    border: 1px solid #007acc;
                }
                QPushButton:pressed {
                    background-color: #2d2d2d;
                }
                QPushButton:disabled {
                    background-color: #2d2d2d;
                    color: #666;
                }
            """)

        return toolbar

    def _create_left_panel(self) -> QWidget:
        group = QGroupBox("文件列表")
        layout = QVBoxLayout(group)

        self.file_table = QTableWidget()
        self.file_table.setColumnCount(4)
        self.file_table.setHorizontalHeaderLabels(["文件名", "类型", "状态", "时间"])
        self.file_table.horizontalHeader().setStretchLastSection(True)
        self.file_table.setAlternatingRowColors(True)
        self.file_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.file_table.itemClicked.connect(self.on_file_selected)
        self.file_table.setStyleSheet("""
            QTableWidget {
                background-color: #252526;
                alternate-background-color: #2d2d2d;
                color: #d4d4d4;
                gridline-color: #3c3c3c;
                border: none;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #333;
                color: #d4d4d4;
                padding: 5px;
                border: 1px solid #3c3c3c;
            }
        """)
        layout.addWidget(self.file_table)

        return group

    def _create_right_panel(self) -> QWidget:
        group = QGroupBox("识别结果")
        layout = QVBoxLayout(group)

        self.result_tabs = QWidget()
        tabs_layout = QVBoxLayout(self.result_tabs)
        tabs_layout.setContentsMargins(0, 0, 0, 0)

        self.raw_text_edit = QTextEdit()
        self.raw_text_edit.setPlaceholderText("识别的文字内容...")
        tabs_layout.addWidget(self.raw_text_edit)

        layout.addWidget(self.result_tabs)

        self.raw_text_edit.setStyleSheet("""
            QTextEdit {
                background-color: #252526;
                color: #d4d4d4;
                border: 1px solid #3c3c3c;
                border-radius: 4px;
                padding: 8px;
                font-family: 'Microsoft YaHei', 'Consolas', monospace;
                font-size: 12px;
            }
        """)

        return group

    def check_models(self):
        models = ModelDetector.detect_all_models()
        all_models = models.get("lmstudio", []) + models.get("ollama", [])
        if not all_models:
            QMessageBox.warning(
                self,
                "未检测到模型",
                "未在本地检测到可用的大模型服务。\n\n请确保：\n1. LM Studio或Ollama已启动\n2. 已加载模型\n3. API服务正在运行\n\n或者在设置中检查API地址是否正确。"
            )

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "选择文件",
            "",
            "支持的文件 (*.pdf *.jpg *.jpeg *.png *.bmp *.gif *.webp);;所有文件 (*.*)"
        )
        if files:
            self.current_files = files
            self.update_file_table()
            self.btn_start.setEnabled(True)

    def select_directory(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择文件夹")
        if dir_path:
            from core.processors import FileScanner
            self.current_files = FileScanner.scan_directory(dir_path)
            self.update_file_table()
            self.btn_start.setEnabled(len(self.current_files) > 0)

    def update_file_table(self):
        self.file_table.setRowCount(len(self.current_files))
        for i, file_path in enumerate(self.current_files):
            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()

            self.file_table.setItem(i, 0, QTableWidgetItem(file_name))
            self.file_table.setItem(i, 1, QTableWidgetItem("PDF" if file_ext == ".pdf" else "图片"))
            self.file_table.setItem(i, 2, QTableWidgetItem("待处理"))
            self.file_table.setItem(i, 3, QTableWidgetItem(""))

            for col in range(4):
                item = self.file_table.item(i, col)
                if item:
                    item.setBackground(QColor("#252526"))

    def start_processing(self):
        if not self.current_files:
            return

        # 处理前检查模型能力
        can_proceed, error_msg = self.processor.can_process()
        if not can_proceed:
            QMessageBox.warning(
                self,
                "无法处理",
                error_msg + "\n\n请在设置中选择正确的视觉模型"
            )
            self.open_settings()
            return

        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)

        self.worker = threading.Thread(target=self._process_worker, daemon=True)
        self.worker.start()

    def _process_worker(self):
        results = []

        def progress_callback(msg):
            self.worker_signals.progress.emit(msg)

        for i, file_path in enumerate(self.current_files):
            result = self.processor.process_file(file_path, progress_callback)
            results.append(result)
            progress = int((i + 1) / len(self.current_files) * 100)
            self.worker_signals.progress.emit(f"进度: {progress}%")

        self.worker_signals.finished.emit(results)

    def on_processing_progress(self, msg: str):
        self.status_label.setText(msg)
        if "进度:" in msg:
            try:
                percent = int(msg.split("进度:")[1].replace("%", "").strip())
                self.progress_bar.setValue(percent)
            except:
                pass

    def on_processing_finished(self, results: List[ProcessingResult]):
        self.results = results
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.progress_bar.setVisible(False)

        success_count = sum(1 for r in results if r.success)
        self.status_label.setText(f"处理完成: 成功 {success_count}/{len(results)}")

        self.update_file_table_results()

        if config.get("auto_export_md") or config.get("auto_export_excel"):
            self.auto_export()

    def update_file_table_results(self):
        for i, result in enumerate(self.results):
            if i >= self.file_table.rowCount():
                break

            status_item = self.file_table.item(i, 2)
            time_item = self.file_table.item(i, 3)

            if result.success:
                status_item.setText("✅ 成功")
                status_item.setForeground(QColor("#4ec9b0"))
            else:
                status_item.setText(f"❌ {result.error[:10]}")
                status_item.setForeground(QColor("#f14c4c"))

            time_item.setText(datetime.now().strftime("%H:%M:%S"))

    def on_file_selected(self, item):
        row = item.row()
        if row < len(self.results):
            result = self.results[row]
            content = f"【文件】{result.file_name}\n\n"
            if result.raw_text:
                content += f"【文字内容】\n{result.raw_text}\n\n"
            if result.summary:
                content += f"【内容摘要】\n{result.summary}\n\n"
            if result.keywords:
                content += f"【关键词】\n{result.keywords}\n\n"
            if result.error:
                content += f"【错误】{result.error}"

            self.raw_text_edit.setPlainText(content)

    def stop_processing(self):
        self.processor.stop()
        self.status_label.setText("已停止处理")
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)

    def export_md(self):
        output_dir = config.output_dir or QFileDialog.getExistingDirectory(
            self, "选择导出目录"
        )
        if not output_dir:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        md_file = os.path.join(output_dir, f"识别结果_{timestamp}.md")

        content = f"# 文档识别结果\n\n"
        content += f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        content += f"使用模型: {config.selected_model}\n\n"
        content += "---\n\n"

        docs = db.get_all()
        for doc in docs:
            content += f"## {doc.file_name}\n\n"
            content += f"- 类型: {doc.file_type}\n"
            content += f"- 处理时间: {doc.processed_at}\n\n"
            if doc.raw_text:
                content += f"### 文字内容\n\n{doc.raw_text}\n\n"
            if doc.summary:
                content += f"### 内容摘要\n\n{doc.summary}\n\n"
            if doc.keywords:
                content += f"### 关键词\n\n{doc.keywords}\n\n"
            content += "---\n\n"

        with open(md_file, "w", encoding="utf-8") as f:
            f.write(content)

        QMessageBox.information(self, "导出成功", f"MD文件已导出到:\n{md_file}")

    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.warning(self, "错误", "请安装openpyxl库: pip install openpyxl")
            return

        output_dir = config.output_dir or QFileDialog.getExistingDirectory(
            self, "选择导出目录"
        )
        if not output_dir:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(output_dir, f"识别结果_{timestamp}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "识别结果"

        headers = ["文件名", "类型", "文字内容", "内容摘要", "关键词", "处理时间", "使用模型"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="007ACC")
            cell.alignment = Alignment(horizontal="center")

        docs = db.get_all()
        for row, doc in enumerate(docs, 2):
            ws.cell(row, 1, doc.file_name)
            ws.cell(row, 2, doc.file_type)
            ws.cell(row, 3, doc.raw_text[:10000])
            ws.cell(row, 4, doc.summary)
            ws.cell(row, 5, doc.keywords)
            ws.cell(row, 6, str(doc.processed_at) if doc.processed_at else "")
            ws.cell(row, 7, doc.model_used)

        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        wb.save(excel_file)
        QMessageBox.information(self, "导出成功", f"Excel文件已导出到:\n{excel_file}")

    def auto_export(self):
        docs = db.get_all()
        if not docs:
            return

        output_dir = config.output_dir
        if not output_dir:
            return

        if config.get("auto_export_md"):
            md_file = os.path.join(output_dir, "识别结果.md")
            content = f"# 文档识别结果\n\n"
            for doc in docs:
                content += f"## {doc.file_name}\n\n"
                if doc.raw_text:
                    content += f"### 文字内容\n\n{doc.raw_text}\n\n"
                if doc.summary:
                    content += f"### 内容摘要\n\n{doc.summary}\n\n"
                if doc.keywords:
                    content += f"### 关键词\n\n{doc.keywords}\n\n"
            with open(md_file, "w", encoding="utf-8") as f:
                f.write(content)

    def open_settings(self):
        dialog = SettingsDialog(self)
        if dialog.exec():
            self.check_models()
